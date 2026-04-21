import argparse
import datetime
import json
import mimetypes
import os
import uuid
from pathlib import Path
from typing import Dict, List, Optional
import hashlib

import pandas as pd
from llama_index.core import SimpleDirectoryReader
from llama_index.core.ingestion import IngestionPipeline
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.readers.base import BaseReader
from llama_index.core.schema import BaseNode, Document, TextNode
from llama_index.readers.file import UnstructuredReader
from llama_index.core.schema import NodeRelationship, RelatedNodeInfo
from openpyxl import load_workbook
from improved_html_cleaner import HtmlCleaner

from chatdku.config import config


def _safe_lower(x):
    return x.strip().lower() if isinstance(x, str) else None


def _import_data(nodes_path: str) -> list:
    with open(nodes_path, "r") as f:
        datas = json.load(f)
    return [TextNode.from_dict(data) for data in datas]


def _ensure_permission_metadata(
    metadata: dict,
    *,
    user_id: str,
    access_type: str | None = None,
    role: str | None = None,
    organization: str | None = None,
) -> dict:
    """Ensure required permission fields exist in node metadata.

    This function sets minimal defaults for ordinary (non-event) nodes.

    Required:
      - access_type: defaults to 'student'
    Optional:
      - role: defaults to 'student' for access_type=='student'
      - organization: None
      - user_id: already present in current metadata, but we keep it consistent
    """
    md = dict(metadata or {})
    if access_type:
        md["access_type"] = access_type.strip().lower()
    else:
        md["access_type"] = md.get("access_type", "student")

    if md["access_type"] == "student":
        md["role"] = _safe_lower(role) or md.get("role") or "student"

    elif md["access_type"] == "office":
        org = _safe_lower(organization) or md.get("organization")
        if not org:
            raise ValueError("organization is required when access_type == 'office'")
        md["organization"] = org

    elif md["access_type"] == "private":
        md["user_id"] = _safe_lower(user_id)

    # fallback
    md.setdefault("role", "student")
    md.setdefault("organization", None)
    md.setdefault("user_id", _safe_lower(user_id))

    return md


def infer_permissions_from_path(file_path: str) -> dict:
    """
    Automatically infer access permissions based on file path.

    Args:
        file_path: Absolute path to the file

    Returns:
        Dictionary with access_type, role, organization, user_id
    """
    path_str = str(file_path)

    # Rule 1: Public knowledge base (visible to everyone)
    if "chat_dku_public" in path_str:
        return {
            "access_type": "public",
            "role": None,
            "organization": None,
            "user_id": None,
        }

    # Rule 2: Main advising knowledge base (visible to students)
    if "chat_dku_advising" in path_str and "test" not in path_str:
        return {
            "access_type": "student",
            "role": "student",
            "organization": None,
            "user_id": None,
        }

    # Rule 3: Test knowledge base (for development/testing)
    if "chat_dku_advising_test" in path_str:
        return {
            "access_type": "student",
            "role": "student",
            "organization": None,
            "user_id": None,
        }

    # Rule 4: Developer personal folders (private, owner-only access)
    developer_names = [
        "Mil3sYu",
        "theta-lin",
        "zhiwei531",
        "Glitterccc",
        "Ar-temis",
        "AndyLu666",
        "SHAO-Jiaqi757",
        "Ederich013",
        "Ada0116",
        "Wangshengyang2004",
        "munishlohani",
        "pomegranar",
        "hafikhan11",
        "GihoonE",
        "algernon-echo",
        "BESTTOOLBOX",
        "sean-allen-siegfreid",
        "zhangyunzhen2027",
    ]

    for name in developer_names:
        if f"/{name}/" in path_str:
            return {
                "access_type": "private",
                "role": None,
                "organization": None,
                "user_id": name,
            }

    # Default: student access (least privilege)
    return {
        "access_type": "student",
        "role": "student",
        "organization": None,
        "user_id": "Chat_DKU",
    }


def load_event_files():
    event_dir = config.event_path
    result = []
    for root, _, files in os.walk(event_dir):
        for f in files:
            path = os.path.join(root, f)
            result.append(path)
    return result


def _write_data(nodes_path: str, data: list):
    with open(nodes_path, "w") as f:
        json.dump(data, f)


def nodes_to_dicts(nodes: list) -> dict:
    result = {
        "ids": [],
        "texts": [],
        "metadatas": [],
    }
    for node in nodes:
        result["ids"].append(node.node_id)
        result["texts"].append(node.text)
        result["metadatas"].append(node.metadata)

    return result


def custom_metadata(user_id: str):
    def _get_meta(file_path: str) -> dict:
        stat = os.stat(file_path)
        return {
            "file_path": file_path,
            "page_number": "Not given.",
            "file_name": os.path.basename(file_path),
            "file_type": (
                mimetypes.guess_type(file_path)[0] or "application/octet-stream"
            ),
            "file_size": stat.st_size,
            "creation_date": datetime.datetime.fromtimestamp(
                stat.st_ctime, datetime.UTC
            ).isoformat(),
            "last_modified_date": datetime.datetime.fromtimestamp(
                stat.st_mtime, datetime.UTC
            ).isoformat(),
            "last_accessed_date": datetime.datetime.fromtimestamp(
                stat.st_atime, datetime.UTC
            ).isoformat(),
            "user_id": user_id,
            "chunk_id": "Not given",
        }

    return _get_meta


class XlsxReader(BaseReader):
    def __init__(
        self,
    ) -> None:
        super().__init__()

    def xlsx_load(self, file: Path) -> str:
        wb = load_workbook(file)
        # Get all sheet names
        sheet_names = wb.sheetnames

        markdown_menu = ""

        # Iterate through each sheet
        for sheet_name in sheet_names:
            sub_wb = wb[sheet_name]
            merged_cells = list(sub_wb.merged_cells.ranges)

            # Iterate through each merged cell
            for merged_cell in merged_cells:
                min_row, max_row = merged_cell.min_row, merged_cell.max_row
                min_col, max_col = merged_cell.min_col, merged_cell.max_col

                # Get the value of the merged cell
                cell_value = sub_wb.cell(row=min_row, column=min_col).value

                # Unmerge the cell
                sub_wb.unmerge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )

                # Fill the value to all cells in the merged range
                for col in range(min_col, max_col + 1):
                    for row in range(min_row, max_row + 1):
                        sub_wb.cell(row=row, column=col, value=cell_value)

            data = list(sub_wb.values)
            header_rows = 1
            if len(data) > 1 and any(data[1]):
                header_rows = 2

            if header_rows == 2:
                combined_header = []
                for h1, h2 in zip(data[0], data[1]):
                    if h1 and h2:
                        combined_header.append(f"{h1.strip()} - {h2.strip()}")
                    elif h1:
                        combined_header.append(h1.strip())
                    elif h2:
                        combined_header.append(h2.strip())
                    else:
                        combined_header.append("Unnamed")
                df = pd.DataFrame(data[2:], columns=combined_header)
            else:
                columns = [c if c else "Unnamed" for c in data[0]]
                df = pd.DataFrame(data[1:], columns=columns)

            # Handle newlines in DataFrame
            df = df.map(
                lambda x: str(x).replace("\n", " ") if isinstance(x, str) else x
            )

            # Drop rows and columns that are completely empty
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)

            # Drop rows and columns that are completely None
            df = df.map(lambda x: None if x == "None" else x)
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)

            structured_rows = []
            for _, row in df.iterrows():
                row_text = "; ".join(
                    [f"{col}: {row[col]}" for col in df.columns if pd.notna(row[col])]
                )
                structured_rows.append(row_text)

            markdown_menu += (
                f"Work Sheet {sheet_name}:\n" + "\n".join(structured_rows) + "\n\n"
            )

        return markdown_menu

    def load_data(
        self, file: Path, extra_info: Optional[Dict] = None
    ) -> List[Document]:
        metadata = {
            "file_name": file.name,
            "file_path": str(file),
        }
        if extra_info is not None:
            metadata.update(extra_info)

        return [Document(text=self.xlsx_load(file), metadata=metadata or {})]


def write_changes(data_dir: str, added_files: set[str], removed_files: set[str]):
    """
    Write the changes that has happened to the log.json.
    Args:
        data_dir: The directory that log.json and all the data is in.
        added_files: Set of added files, that was returned from read_changes()
        removed_files: Set of removed files, that was returned from read_changes()
    """
    log_path = os.path.join(data_dir, "log.json")

    with open(log_path, "r") as file:
        log = json.load(file)

    new_list = []
    for f in log["file_paths"]:
        # filtering event out
        if f.startswith(config.event_path):
            continue
        if f not in removed_files:
            new_list.append(f)

    # filtering event out
    for f in added_files:
        if not f.startswith(config.event_path):
            new_list.append(f)

    log["file_paths"] = new_list

    with open(log_path, "w") as file:
        json.dump(log, file)


def read_changes(data_dir: str) -> tuple[set[str], set[str]]:
    """
    Read the log.json file and read which files are turned into nodes.
    Will skip files with suffixes ".json", and "pkl".
    Args:
        data_dir: The directory that log.json and all the data is in.
    Returns:
        tuple(added_files, removed_files): A tuple of sets with the added files and removed files.
    """
    log_path = os.path.join(data_dir, "log.json")

    # Load previous file paths from log
    if os.path.exists(log_path):
        with open(log_path, "r") as file:
            log = json.load(file)
            previous_files = [
                f
                for f in log.get("file_paths", [])
                if not f.startswith(config.event_path)
            ]
    else:
        previous_files = []
        with open(log_path, "w") as file:
            json.dump({"file_paths": []}, file)

    # Recursively get current file paths
    current_files = []
    for root, _, files in os.walk(data_dir):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            # Skip log.json itself
            if (
                os.path.abspath(file_path) == os.path.abspath(log_path)
                or file_name.endswith(".json")
                or file_name.endswith(".pkl")
                or file_path.startswith(os.path.abspath(config.event_path))
            ):
                continue
            current_files.append(os.path.abspath(file_path))

    # Compute added and removed files
    added_files = set(current_files) - set(previous_files)
    removed_files = set(previous_files) - set(current_files)

    return added_files, removed_files


def clean_expired_nodes(nodes):
    now = datetime.datetime.now(datetime.timezone.utc)

    kept = []
    for n in nodes:
        meta = n.metadata
        if meta.get("is_event") is True:
            exp = meta.get("expire_at")
            if exp:
                exp_dt = datetime.datetime.fromisoformat(exp.replace("Z", "+00:00"))
                if exp_dt < now:
                    # expired -> do not keep
                    continue
        kept.append(n)

    return kept


def _read_pdf(
    file_paths: list[str], user_id, access_type, role, organization
) -> list[TextNode]:
    """
    Read PDF files using LlamaParse + fixed chunk size SentenceSplitter.

    This is the rollback version:
    - remove structure-aware chunking
    - restore fixed chunk size splitting
    """
    from llama_parse import LlamaParse

    total_nodes = []

    for file_path in file_paths:
        print(f"Reading PDF with LlamaParse: {file_path}")

        parser = LlamaParse(
            api_key=config.llamaparse_api,
            result_type="markdown",
            num_workers=4,
            verbose=True,
            language="en",
            parsing_instruction="This is a DKU document. Extract all text and preserve structure.",
        )

        documents = parser.load_data(file_path)

        splitter = SentenceSplitter(
            chunk_size=config.chunk_size,
            chunk_overlap=config.chunk_overlap,
        )
        nodes = splitter.get_nodes_from_documents(documents)

        for node in nodes:
            chunk_id = str(uuid.uuid4())

            base_metadata = custom_metadata(user_id)(file_path)
            base_metadata["chunk_id"] = chunk_id
            base_metadata["chunking_method"] = "fixed"

            base_metadata = _ensure_permission_metadata(
                base_metadata,
                user_id=user_id,
                access_type=access_type,
                role=role,
                organization=organization,
            )

            node.id_ = chunk_id
            node.metadata = base_metadata

            doc_id = os.path.abspath(file_path)
            node.relationships[NodeRelationship.SOURCE] = RelatedNodeInfo(
                node_id=hashlib.md5(doc_id.encode()).hexdigest()
            )

            total_nodes.append(node)

        print(f"Finished loading {file_path}. Generated {len(nodes)} chunks.")

    return total_nodes


def _read_non_pdf(
    files: list, user_id, access_type, role, organization
) -> list[BaseNode]:
    """
    Read non-PDF files using UnstructuredReader and other readers.

    Args:
        files: List of file paths to process
        user_id: User ID for metadata
        access_type: Access type (public/student/office/private)
        role: User role
        organization: Organization name (required for office access)

    Returns:
        List of BaseNode objects
    """

    # Allow callers to pass an empty list (e.g., event folder is empty)
    if not files:
        return []
    reader = UnstructuredReader()
    xlsx_reader = XlsxReader()
    html_cleaner = HtmlCleaner()
    non_pdf_documents = SimpleDirectoryReader(
        input_files=files,
        file_metadata=custom_metadata(user_id),
        required_exts=[".htm", ".html", ".csv", ".jpg", ".xlsx"],
        file_extractor={
            ".htm": html_cleaner,
            ".html": html_cleaner,
            ".csv": reader,
            ".jpg": reader,
            ".xlsx": xlsx_reader,
        },
    ).load_data(show_progress=True)

    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(
                chunk_size=config.chunk_size,
                chunk_overlap=config.chunk_overlap,
            ),
        ]
    )

    non_pdf_nodes = pipeline.run(documents=non_pdf_documents, show_progress=True)

    for node in non_pdf_nodes:
        if node.text == "":
            continue
        node.node_id = str(uuid.uuid4())
        node.metadata["chunk_id"] = node.node_id

        # Permission defaults for ordinary nodes
        node.metadata = _ensure_permission_metadata(
            node.metadata,
            user_id=user_id,
            access_type=access_type,
            role=role,
            organization=organization,
        )

        file_path = (
            node.metadata.get("file_path")
            or node.metadata.get("source")
            or node.metadata.get("file_name")
        )
        if not file_path:
            raise ValueError("Cannot determine file_path for node")
        doc_id = os.path.abspath(file_path) if file_path else "unknown"
        node.relationships[NodeRelationship.SOURCE] = RelatedNodeInfo(
            node_id=hashlib.md5(doc_id.encode()).hexdigest()
        )

        if str(node.metadata.get("file_name", "")).endswith(".pptx"):
            node.metadata["extraction_errors"] = str(
                node.metadata.get("extraction_errors")
            )
            node.metadata["extraction_warnings"] = str(
                node.metadata["extraction_warnings"]
            )
            node.metadata["tables"] = str(node.metadata["tables"])
            node.metadata["charts"] = str(node.metadata["charts"])
            node.metadata["images"] = str(node.metadata["images"])
            node.metadata["text_sections"] = str(node.metadata["text_sections"])

    return non_pdf_nodes


def update_events(user_id: str) -> list:
    """Always re-read all event files and generate fresh event nodes."""
    event_files = load_event_files()

    # If the event folder is missing or empty, just skip event ingestion.
    if not event_files:
        return []

    event_nodes = _read_non_pdf(
        event_files,
        user_id,
        access_type="student",
        role="student",
        organization=None,
    )

    now = datetime.datetime.now(datetime.timezone.utc)
    expire = now + datetime.timedelta(days=7)

    for n in event_nodes:
        n.metadata["is_event"] = True
        n.metadata["expire_at"] = expire.isoformat().replace("+00:00", "Z")

    return event_nodes


def update(
    data_dir: str,
    user_id: str,
    access_type: str = None,
    role: str = None,
    organization: str = None,
    verbose: bool = False,
):
    """
    Main update function that processes all documents and generates nodes.

    This function:
    1. Detects added/removed files since last run
    2. Processes new PDF files using LlamaParse + fixed chunk size
    3. Processes new non-PDF files using UnstructuredReader
    4. Removes nodes from deleted files
    5. Generates fresh event nodes
    6. Saves all nodes to nodes.json

    Args:
        data_dir: Root directory containing documents
        user_id: User ID for metadata
        access_type: Access type (public/student/office/private). If None, auto-infer from path
        role: User role. If None, auto-infer from path
        organization: Organization name (required for office access)
        verbose: Whether to print detailed information
    """
    # detect add/remove only in NON-EVENT dir
    nodes_path = os.path.join(data_dir, "nodes.json")

    # load old nodes
    if os.path.exists(nodes_path):
        old_nodes = _import_data(nodes_path)
    else:
        old_nodes = []

    # keep only non-event nodes
    non_event_old = [n for n in old_nodes if not n.metadata.get("is_event")]
    new_nodes = []

    added_files, removed_files = read_changes(data_dir)
    if verbose:
        print(f"Files to be added: {added_files}")
        print(f"Files to be removed: {removed_files}")

    # load newly added non-event files with per-file permission inference
    if added_files:
        pdf_files = [file for file in added_files if file.endswith(".pdf")]
        non_pdf_files = list(set(added_files) - set(pdf_files))

        # Process non-PDF files with per-file permission inference
        for file in non_pdf_files:
            # Determine permissions for this file
            if access_type:
                # Global override mode: use provided parameters
                file_access_type = access_type
                file_role = role
                file_organization = organization
            else:
                # Auto-inference mode: determine from file path
                perms = infer_permissions_from_path(file)
                file_access_type = perms["access_type"]
                file_role = perms["role"]
                file_organization = perms["organization"]

            # Process single file
            nodes = _read_non_pdf(
                [file],
                user_id,
                file_access_type,
                file_role,
                file_organization,
            )
            new_nodes.extend(nodes)
            if verbose:
                print(
                    f"Processed {file} -> access_type={file_access_type}, role={file_role}"
                )

        # Process PDF files with per-file permission inference
        for file in pdf_files:
            # Determine permissions for this file
            if access_type:
                # Global override mode: use provided parameters
                file_access_type = access_type
                file_role = role
                file_organization = organization
            else:
                # Auto-inference mode: determine from file path
                perms = infer_permissions_from_path(file)
                file_access_type = perms["access_type"]
                file_role = perms["role"]
                file_organization = perms["organization"]

            # Process single file
            nodes = _read_pdf(
                [file],
                user_id,
                file_access_type,
                file_role,
                file_organization,
            )
            new_nodes.extend(nodes)
            if verbose:
                print(
                    f"Processed {file} -> access_type={file_access_type}, role={file_role}"
                )

        print("Total added nodes:", len(new_nodes))

    # remove deleted non-event files
    kept_nodes = []
    for n in non_event_old:
        file_path_meta = os.path.abspath(n.metadata.get("file_path", ""))
        if file_path_meta not in removed_files:
            kept_nodes.append(n)
    print("Total kept nodes:", len(kept_nodes))

    # combine non-event
    non_event_nodes = kept_nodes + new_nodes
    # now load events
    event_nodes = update_events(user_id)

    total_nodes = non_event_nodes + event_nodes
    print("Total nodes:", len(total_nodes))

    nodes_dicts = [node.to_dict() for node in total_nodes]

    _write_data(nodes_path, nodes_dicts)
    write_changes(data_dir, added_files, removed_files)

    print("Document load done!")


def main(data_dir, user_id, access_type, role, organization=None, verbose=False):
    """
    Main entry point for the document processing script.

    Args:
        data_dir: Directory containing documents to process
        user_id: User ID for metadata
        access_type: Access type (public/student/office/private)
        role: User role
        organization: Organization name (required when access_type == 'office')
        verbose: Whether to print detailed information
    """
    if data_dir is None:
        data_dir = config.data_dir
    if user_id is None:
        user_id = "Chat_DKU"
    if access_type is None:
        access_type = "student"
    if role is None:
        role = "student"

    update(data_dir, user_id, access_type, role, organization, verbose=verbose)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process data directory path")
    parser.add_argument(
        "--data_dir",
        type=str,
        default=config.data_dir,
        help="The directory containing the data",
    )
    parser.add_argument(
        "--user_id",
        type=str,
        default="Chat_DKU",
        help="ID of the user. Defaults to Chat_DKU if none given.",
    )
    parser.add_argument(
        "--access_type",
        type=str,
        default="student",
        help="Access type for the nodes. Including 'public', 'student', 'office', 'private'. Defaults to 'student'.",
    )
    parser.add_argument(
        "--role",
        type=str,
        default="student",
        help="Role for the nodes. Defaults to 'student'.",
    )
    parser.add_argument(
        "--organization",
        type=str,
        default=None,
        help="Organization for the nodes. Required when access_type == 'office'.",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        default=True,
        help="Whether to print extra information.",
    )
    args = parser.parse_args()

    main(
        args.data_dir,
        args.user_id,
        args.access_type,
        args.role,
        args.organization,
        args.verbose,
    )
