import os
import nltk
import nest_asyncio

nest_asyncio.apply()

import pickle
import argparse
import json
from llama_index.core.readers.base import BaseReader
from llama_index.readers.file import UnstructuredReader
from chatdku.config import config
from markdownify import markdownify as md
from tqdm import tqdm
from pathlib import Path
from typing import Any, Dict, List, Optional
from llama_index.core.schema import Document, TextNode
import pandas as pd
from openpyxl import load_workbook
import hashlib

# Import structure-aware PDF chunker (local parsing, replaces LlamaParse)
from structure_chunker import process_pdf as process_pdf_structure_aware


# Override detect_filetype so that html files containing JavaScript code are loaded in html format.
import unstructured.file_utils.filetype
from custom_filetype_detect import custom_detect_filetype

unstructured.file_utils.filetype.detect_filetype = custom_detect_filetype

# Override auto partation
import unstructured.partition.auto
from custom_partation import partition

unstructured.partition.auto.partition = partition


class XlsxReader(BaseReader):
    def __init__(
        self,
    ) -> None:
        super().__init__()
    
    
    def xlsx_load(self,file: Path) -> str:
        wb = load_workbook(file)
        # 获取所有工作表的名称
        sheet_names = wb.sheetnames

        markdown_menu = ""

        # 遍历每一个工作表
        for sheet_name in sheet_names:
            sub_wb = wb[sheet_name]
            merged_cells = list(sub_wb.merged_cells.ranges)  # 转换为列表
            
            # 遍历每一个合并单元格
            for merged_cell in merged_cells:
                min_row, max_row = merged_cell.min_row, merged_cell.max_row
                min_col, max_col = merged_cell.min_col, merged_cell.max_col
                
                # 获取合并单元格的值
                cell_value = sub_wb.cell(row=min_row, column=min_col).value
                
                # 解除合并单元格
                sub_wb.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                
                # 将值填充到之前合并单元格的所有单元格中
                for col in range(min_col, max_col + 1):
                    for row in range(min_row, max_row + 1):
                        sub_wb.cell(row=row, column=col, value=cell_value)

            data = wb[sheet_name].values
            columns = next(data)[0:]  # 获取第一行作为列名
            df = pd.DataFrame(data, columns=columns)
            
            # 处理 DataFrame 中的回车符
            df = df.applymap(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)
            
            # 去掉全为空值的行和列
            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # 去掉全为 None 的行和列
            df = df.applymap(lambda x: None if x == 'None' else x)
            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # 将 DataFrame 转换为 Markdown 格式
            markdown_output = df.to_markdown(index=False)
            markdown_menu += sheet_name + "菜单：\n"
            markdown_menu += markdown_output
            markdown_menu += "\n\n\n\n"

        return markdown_menu


    def load_data(
        self, file: Path, extra_info: Optional[Dict] = None
    ) -> List[Document]:
        docs = []
        metadata = {
            "file_name": file.name,
            "file_path": str(file),
        }
        if extra_info is not None:
            metadata.update(extra_info)

        return [Document(text=self.xlsx_load(file), metadata=metadata or {})]




def hash_file(filename):
    h = hashlib.sha256()
    with open(filename, "rb") as file:
        while True:
            chunk = file.read(h.block_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def hash_directory(directory):
    all_hashes = ""
    for root, _, files in os.walk(directory):
        for filename in files:
            filepath = os.path.join(root, filename)
            file_hash = hash_file(filepath)
            all_hashes += file_hash
    final_hash = hashlib.sha256(all_hashes.encode("utf-8")).hexdigest()
    return final_hash


def update_data(data_dir):
    # Required for UnstructuredReader
    # nltk.download("averaged_perceptron_tagger")
    reader = UnstructuredReader()

    documents_path = "/home/Glitterccc/ChatDKU/documents/menu_document.pkl"
    documents_path = "/home/Glitterccc/ChatDKU/documents/menu_document.pkl"

    reader = UnstructuredReader()
    xlsx_reader = XlsxReader()
    
    all_documents = []
    
    # Walk through directory and process files by type
    for root, dirs, files in os.walk(data_dir):
        for file in files:
            file_path = os.path.join(root, file)
            file_ext = os.path.splitext(file)[1].lower()
            
            try:
                # PDF files: use structure_chunker (local parsing)
                if file_ext == '.pdf':
                    chunked_docs = process_pdf_structure_aware(
                        file_path,
                        max_chunk_size=800,
                        min_chunk_size=80
                    )
                    all_documents.extend(chunked_docs)
                
                # HTML/HTM files
                elif file_ext in ['.html', '.htm']:
                    docs = reader.load_data(file_path)
                    for doc in docs:
                        # 处理文档
                        if doc.metadata.get("file_type") == "text/html":
                            with open(file_path, "r") as f:
                                html = f.read()
                            try:
                                doc.text = md(html)
                            except Exception as e:
                                # Fallback: keep original text if markdown conversion fails
                                print(f"Warning: failed to convert HTML to markdown for {file_path}: {e}")
                        all_documents.append(doc)
                
                # XLSX files
                elif file_ext == '.xlsx':
                    docs = xlsx_reader.load_data(Path(file_path))
                    all_documents.extend(docs)
                
                # CSV files
                elif file_ext == '.csv':
                    docs = reader.load_data(file_path)
                    all_documents.extend(docs)
                
                # Image files
                elif file_ext in ['.jpg', '.jpeg', '.png']:
                    docs = reader.load_data(file_path)
                    all_documents.extend(docs)
                
                # Other supported file types
                elif file_ext in ['.txt', '.docx', '.doc']:
                    docs = reader.load_data(file_path)
                    all_documents.extend(docs)
                    
            except Exception:
                continue
    
    # Convert Document to TextNode for compatibility with load_redis/load_chroma/load_postgres
    nodes = []
    for doc in all_documents:
        node = TextNode(
            text=doc.text,
            metadata=doc.metadata,
            id_=doc.id_ if hasattr(doc, 'id_') else None
        )
        nodes.append(node)
    
    # Save as nodes.json (used by downstream loaders)
    nodes_path = config.nodes_path if hasattr(config, 'nodes_path') else os.path.join(data_dir, "nodes.json")
    os.makedirs(os.path.dirname(nodes_path), exist_ok=True)
    
    with open(nodes_path, 'w') as f:
        json.dump([node.to_dict() for node in nodes], f, indent=2, default=str)
    
    # Also save as documents.pkl for backward compatibility
    with open(documents_path, "wb") as f:
        pickle.dump(all_documents, f)
    
    return all_documents


def main(data_dir=None):
    if data_dir is None:
        data_dir = config.data_dir

    update_data(data_dir)
    hash = hash_directory(data_dir)
    hash_path = os.path.join("./", "hash.pkl")
    with open(hash_path, "wb") as hf:
        pickle.dump(hash, hf)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process data directory path")
    parser.add_argument("data_dir", type=str, help="The directory containing the data")
    args = parser.parse_args()

    main(args.data_dir)
