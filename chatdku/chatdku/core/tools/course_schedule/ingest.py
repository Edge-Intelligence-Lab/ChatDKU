#!/usr/bin/env python3

import argparse
import os
import sys
import json
import logging
import hashlib
from pathlib import Path
from datetime import datetime

import psycopg2
from psycopg2 import sql
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
from chatdku.chatdku.config import config

from openpyxl import load_workbook


DEFAULT_DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "database": "chatdku_db",
    "user": "chatdku_user",
}


class ScheduleIngestor:
    """Ingest DKU course schedule data into PostgreSQL"""

    def __init__(self, args):
        self.args = args

        if not args.term or not str(args.term).strip():
            raise ValueError(
                "Argument --term is required and cannot be empty. "
                "Example: --term 'Spring 2026'"
            )

        self.term = args.term.strip()
        self.setup_logging()
        self.setup_database_connection()
        self.ensure_tracking_table()
        self.load_schema()

    def setup_logging(self):
        log_file = Path(self.args.pool) / "schedule_ingestor.log"
        log_file.parent.mkdir(parents=True, exist_ok=True)

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.FileHandler(log_file, mode="a"),
                logging.StreamHandler(sys.stdout),
            ],
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info("Schedule Ingestor started")

    def setup_database_connection(self):
        db_config = DEFAULT_DB_CONFIG.copy()

        if self.args.db_host:
            db_config["host"] = self.args.db_host
        if self.args.db_port:
            db_config["port"] = self.args.db_port
        if self.args.db_name:
            db_config["database"] = self.args.db_name
        if self.args.db_user:
            db_config["user"] = self.args.db_user

        db_config["password"] = os.environ.get("DB_PWD")

        self.conn = psycopg2.connect(**db_config)
        self.conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
        self.logger.info(f"Connected to PostgreSQL database: {db_config['database']}")

    def ensure_tracking_table(self):
        """Create file tracking table if it doesn't exist"""
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ingested_files (
                id SERIAL PRIMARY KEY,
                file_path TEXT NOT NULL UNIQUE,
                file_hash TEXT NOT NULL,
                term TEXT NOT NULL,
                ingested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                row_count INTEGER DEFAULT 0
            )
        """)
        cursor.close()
        self.logger.info("File tracking table ensured")

    def get_file_hash(self, file_path: Path) -> str:
        """Calculate MD5 hash of file for change detection"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def is_file_processed(self, file_path: Path) -> tuple[bool, str]:
        """
        Check if file has been processed before
        Returns: (is_processed, previous_hash)
        """
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT file_hash FROM ingested_files WHERE file_path = %s",
            (str(file_path),)
        )
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            return True, result[0]
        return False, None

    def mark_file_processed(self, file_path: Path, file_hash: str, row_count: int):
        """Record file as processed in tracking table"""
        cursor = self.conn.cursor()
        cursor.execute("""
            INSERT INTO ingested_files (file_path, file_hash, term, row_count)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (file_path) 
            DO UPDATE SET 
                file_hash = EXCLUDED.file_hash,
                term = EXCLUDED.term,
                ingested_at = CURRENT_TIMESTAMP,
                row_count = EXCLUDED.row_count
        """, (str(file_path), file_hash, self.term, row_count))
        cursor.close()

    def delete_old_data_for_file(self, file_path: Path):
        """Delete existing data from a file before re-ingesting"""
        # This is a simple approach - you might want to track by term instead
        # For now, we'll assume re-processing means updating the same term
        cursor = self.conn.cursor()
        cursor.execute(
            "DELETE FROM dku_class_schedule WHERE term = %s",
            (self.term,)
        )
        deleted_count = cursor.rowcount
        cursor.close()
        self.logger.info(f"Deleted {deleted_count} existing rows for term '{self.term}'")

    def load_schema(self):
        schema_path = Path(self.args.schema)
        if not schema_path.is_absolute():
            schema_path = Path(__file__).parent / schema_path

        with open(schema_path, "r") as f:
            self.schema = json.load(f)

        self.logger.info(f"Schema loaded from: {schema_path}")

    def store_row(self, row: dict, table="dku_class_schedule"):
        cursor = self.conn.cursor()
        columns = list(row.keys())
        values = [row[c] for c in columns]

        insert_sql = sql.SQL(
            "INSERT INTO {table} ({fields}) VALUES ({placeholders})"
        ).format(
            table=sql.Identifier(table),
            fields=sql.SQL(", ").join(map(sql.Identifier, columns)),
            placeholders=sql.SQL(", ").join(sql.Placeholder() * len(columns)),
        )

        cursor.execute(insert_sql, values)
        cursor.close()

    # ===== XLSX PART =====

    def process_xlsx(self, file_path: Path):
        self.logger.info(f"Processing schedule XLSX: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        total_rows = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.logger.info(f"Processing sheet: {sheet_name}")

            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # 第二行才是 column name
            headers = [str(h).strip() if h else None for h in rows[1]]

            for row_values in rows[2:]:
                row_dict = {
                    headers[i]: row_values[i]
                    for i in range(len(headers))
                    if headers[i] is not None
                }

                cleaned = self.clean_row(row_dict)
                cleaned["term"] = self.term
                
                if "term" not in cleaned or not cleaned["term"]:
                    raise RuntimeError(
                        f"Missing term when ingesting row from {file_path}"
                    )
                self.store_row(cleaned)
                total_rows += 1

        return total_rows

    # ===== CLEANING =====

    def clean_row(self, row: dict) -> dict:
        EXCEL_TO_DB_COL = {
            # "Term": "term",
            "Session": "session",
            "Subject": "subject",
            "Catalog": "catalog",
            "Section": "section",
            "Name": "instructor_name",
            "Descr": "course_title",
            "Facil ID": "facility_id",
            "Mtg Start": "meeting_start",
            "Mtg End": "meeting_end",
            "Max Units": "max_units",   
            "Type": "class_type",
            "Email": "instructor_email",
            "Preferred": "preferred"
        }

        cleaned = {}

        def empty_to_none(v):
            return None if v is None or v == "" else v

        # 映射列
        for key, value in row.items():
            db_col = EXCEL_TO_DB_COL.get(key)
            if db_col:
                cleaned[db_col] = empty_to_none(value)
        if "catalog" in cleaned and isinstance(cleaned["catalog"], str):
            cleaned["catalog"] = cleaned["catalog"].strip()

        # 处理 days 列，组合 Mon-Fri
        days_list = []
        for day in ["Mon","Tues","Wed","Thurs","Fri"]:
            val = row.get(day)
            if isinstance(val, str) and val.strip().upper() in ("Y", "YES"):
                days_list.append(day)
        if days_list:
            cleaned["days"] = days_list

        # 处理时间
        for k in ["meeting_start", "meeting_end"]:
            if k in cleaned and cleaned[k]:
                if isinstance(cleaned[k], str) and len(cleaned[k]) == 5:  # HH:MM
                    cleaned[k] += ":00"

        # 处理 max_units
        if "max_units" in cleaned and cleaned["max_units"]:
            cleaned["max_units"] = float(cleaned["max_units"])

        # 全部列名小写
        cleaned = {k.lower(): v for k, v in cleaned.items()}
        return cleaned

    def process_pool(self):
        pool_path = Path(self.args.pool)
        self.logger.info(f"Resolved pool path: {pool_path.resolve()}")

        if not pool_path.exists():
            self.logger.error(f"Pool directory does not exist: {pool_path}")
            sys.exit(1)

        xlsx_files = list(pool_path.rglob("*.xlsx"))
        if not xlsx_files:
            self.logger.info("No XLSX schedule files found")
            return

        self.logger.info(f"Found {len(xlsx_files)} XLSX files")

        new_files = []
        updated_files = []
        skipped_files = []

        for file_path in xlsx_files:
            current_hash = self.get_file_hash(file_path)
            is_processed, previous_hash = self.is_file_processed(file_path)

            if is_processed:
                if current_hash == previous_hash:
                    self.logger.info(f"SKIPPED (unchanged): {file_path.name}")
                    skipped_files.append(file_path)
                    continue
                else:
                    self.logger.info(f"UPDATED (file changed): {file_path.name}")
                    updated_files.append(file_path)
                    if self.args.update_mode:
                        # Delete old data before re-ingesting
                        self.delete_old_data_for_file(file_path)
            else:
                self.logger.info(f"NEW file detected: {file_path.name}")
                new_files.append(file_path)

            # Process the file
            row_count = self.process_xlsx(file_path)
            self.mark_file_processed(file_path, current_hash, row_count)
            self.logger.info(f"Ingested {row_count} rows from {file_path.name}")

        # Summary
        self.logger.info("=" * 60)
        self.logger.info("INGESTION SUMMARY")
        self.logger.info("=" * 60)
        self.logger.info(f"New files processed: {len(new_files)}")
        self.logger.info(f"Updated files processed: {len(updated_files)}")
        self.logger.info(f"Unchanged files skipped: {len(skipped_files)}")
        self.logger.info(f"Total files processed: {len(new_files) + len(updated_files)}")
        
        if new_files:
            self.logger.info("\nNew files:")
            for f in new_files:
                self.logger.info(f"  - {f.name}")
        
        if updated_files:
            self.logger.info("\nUpdated files:")
            for f in updated_files:
                self.logger.info(f"  - {f.name}")

        self.logger.info("=" * 60)

    def cleanup(self):
        if hasattr(self, "conn"):
            self.conn.close()
            self.logger.info("Database connection closed")


def main():
    parser = argparse.ArgumentParser(
        description="Ingest DKU course schedule XLSX into PostgreSQL with update tracking"
    )

    parser.add_argument(
        "--pool",
        default=config.course_schedule_path,
        help="Directory containing schedule XLSX files",
    )

    parser.add_argument(
        "--schema",
        default="schedule_schema.json",
        help="JSON schema file",
    )

    parser.add_argument(
        "--term",
        default="Spring 2026",
        help="Term of the course schedule (e.g., Spring 2026)",
    )

    parser.add_argument(
        "--update-mode",
        action="store_true",
        help="Enable update mode: delete old data for the term before re-ingesting",
    )

    parser.add_argument("--db-host")
    parser.add_argument("--db-port", type=int)
    parser.add_argument("--db-name")
    parser.add_argument("--db-user")

    args = parser.parse_args()

    ingestor = None
    try:
        ingestor = ScheduleIngestor(args)
        ingestor.process_pool()
    finally:
        if ingestor:
            ingestor.cleanup()


if __name__ == "__main__":
    main()