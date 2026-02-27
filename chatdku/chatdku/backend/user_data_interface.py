import os
import mimetypes
import datetime
import nest_asyncio
import uuid
import json
import chromadb
from redis import Redis
from redisvl.schema import IndexSchema
from chromadb.utils.embedding_functions import HuggingFaceEmbeddingServer
from llama_index.core import SimpleDirectoryReader
from llama_index.vector_stores.redis import RedisVectorStore
from llama_index.core.schema import TextNode
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.ingestion import IngestionPipeline
from llama_index.core.readers.base import BaseReader
from llama_index.readers.file import UnstructuredReader
from llama_parse import LlamaParse
from chatdku.config import config
from pathlib import Path
from typing import Dict, List, Optional
from llama_index.core.schema import Document
import pandas as pd
from openpyxl import load_workbook


from chatdku.ingestion.load_redis import load_redis

# Override detect_filetype so that html files containing JavaScript code are loaded in html format.
import unstructured.file_utils.filetype
from chatdku.ingestion.custom_filetype_detect import custom_detect_filetype

# Override auto partation
import unstructured.partition.auto
from chatdku.ingestion.custom_partation import partition

nest_asyncio.apply()
unstructured.file_utils.filetype.detect_filetype = custom_detect_filetype

unstructured.partition.auto.partition = partition


def nodes_to_dicts(nodes: list):
    result = {
        "ids": [],
        "texts": [],
        "metadatas": [],
    }
    for node in nodes:
        result["ids"].append(node.node_id)
        result["texts"].append(node.text)
        result["metadatas"].append(node.metadata)

    return result


def custom_metadata(user_id: str):
    def _get_meta(file_path: str) -> dict:
        stat = os.stat(file_path)
        return {
            "file_path": file_path,
            "page_number": "Not given.",
            "file_name": os.path.basename(file_path),
            "file_type": mimetypes.guess_type(file_path)[0]
            or "application/octet-stream",
            "file_size": stat.st_size,
            "creation_date": datetime.datetime.utcfromtimestamp(
                stat.st_ctime
            ).isoformat(),
            "last_modified_date": datetime.datetime.utcfromtimestamp(
                stat.st_mtime
            ).isoformat(),
            "last_accessed_date": datetime.datetime.utcfromtimestamp(
                stat.st_atime
            ).isoformat(),
            "user_id": user_id,
            "chunk_id": "Not given",
        }

    return _get_meta


class XlsxReader(BaseReader):
    def __init__(
        self,
    ) -> None:
        super().__init__()

    def xlsx_load(self, file: Path) -> str:
        wb = load_workbook(file)
        # 获取所有工作表的名称
        sheet_names = wb.sheetnames

        markdown_menu = ""

        # 遍历每一个工作表
        for sheet_name in sheet_names:
            sub_wb = wb[sheet_name]
            merged_cells = list(sub_wb.merged_cells.ranges)  # 转换为列表

            # 遍历每一个合并单元格
            for merged_cell in merged_cells:
                min_row, max_row = merged_cell.min_row, merged_cell.max_row
                min_col, max_col = merged_cell.min_col, merged_cell.max_col

                # 获取合并单元格的值
                cell_value = sub_wb.cell(row=min_row, column=min_col).value

                # 解除合并单元格
                sub_wb.unmerge_cells(
                    start_row=min_row,
                    start_column=min_col,
                    end_row=max_row,
                    end_column=max_col,
                )

                # 将值填充到之前合并单元格的所有单元格中
                for col in range(min_col, max_col + 1):
                    for row in range(min_row, max_row + 1):
                        sub_wb.cell(row=row, column=col, value=cell_value)

            data = wb[sheet_name].values
            columns = next(data)[0:]  # 获取第一行作为列名
            df = pd.DataFrame(data, columns=columns)

            # 处理 DataFrame 中的回车符
            df = df.applymap(
                lambda x: str(x).replace("\n", " ") if isinstance(x, str) else x
            )

            # 去掉全为空值的行和列
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)

            # 去掉全为 None 的行和列
            df = df.applymap(lambda x: None if x == "None" else x)
            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)

            # 将 DataFrame 转换为 Markdown 格式
            markdown_output = df.to_markdown(index=False)
            markdown_menu += sheet_name + "菜单：\n"
            markdown_menu += markdown_output
            markdown_menu += "\n\n\n\n"

        return markdown_menu

    def load_data(
        self, file: Path, extra_info: Optional[Dict] = None
    ) -> List[Document]:
        docs = []
        metadata = {
            "file_name": file.name,
            "file_path": str(file),
        }
        if extra_info is not None:
            metadata.update(extra_info)

        return [Document(text=self.xlsx_load(file), metadata=metadata or {})]


def write_changes(data_dir: str, new_files: list[Dict], removed_files: list[str]):
    log_path = os.path.join(data_dir, "log.json")

    with open(log_path, "r") as file:
        log = json.load(file)

    for file in log["file_paths"][:]:
        if file in removed_files:
            log["file_paths"].remove(file)

    log["file_paths"].extend(new_files)

    with open(log_path, "w") as file:
        json.dump(log, file)


def read_changes(data_dir: str):
    log_path = os.path.join(data_dir, "log.json")

    if os.path.exists(log_path):
        with open(log_path, "r") as file:
            log = json.load(file)
            previous_files = log.get("file_paths", [])
    else:
        previous_files = []
        with open(log_path, "w") as file:
            json.dump({"file_paths": []}, file)

    current_files = []
    for file_name in os.listdir(data_dir):
        if file_name.endswith(".json"):
            continue
        file_path = os.path.join(data_dir, file_name)
        current_files.append(file_path)

    removed_files = list(set(previous_files) - set(current_files))

    added_files = list(set(current_files) - set(previous_files))

    return added_files, removed_files


# For large files doing one collection.add seems to break stuff.
def embed_pdf(file_paths: list[str], user_id, collection):
    total_nodes = []
    llama_parse_api_key = os.getenv("LLAMA_PARSE_API_KEY")
    if not llama_parse_api_key:
        raise RuntimeError(
            "LLAMA_PARSE_API_KEY is not set. Set it to enable PDF parsing."
        )
    parser = SentenceSplitter(
        chunk_size=1024,
        chunk_overlap=20,
    )
    pdf_reader = LlamaParse(
        api_key=llama_parse_api_key,
        result_type="json",
        verbose=True,
        disable_image_extraction=True,
    )
    for file_path in file_paths:
        print(f"Reading: {file_path}")  # Debugging output
        if not os.path.exists(file_path):
            print(f"Error: The directory '{file_path}' does not exist.")
        pdf_loader = pdf_reader.parse(file_path)
        nodes_buffer = []

        for i, page in enumerate(pdf_loader.pages):
            metadata = custom_metadata(user_id)(file_path)
            # Adding page number
            metadata["page_number"] = page.page

            chunks = parser.split_text(page.md)
            for chunk in chunks:
                chunk_id = str(uuid.uuid4())

                metadata["chunk_id"] = chunk_id

                node = TextNode(
                    text=chunk,
                    id_=chunk_id,
                    metadata=metadata,
                )
                nodes_buffer.append(node)
                total_nodes.append(node)

            # for every 25 pages we upload them to chroma
            if i % 25 == 0:
                nodes_buffer_dict = nodes_to_dicts(nodes_buffer)
                collection.add(
                    ids=nodes_buffer_dict["ids"],
                    documents=nodes_buffer_dict["texts"],
                    metadatas=nodes_buffer_dict["metadatas"],
                )
                nodes_buffer = []

        if nodes_buffer:
            nodes_buffer_dict = nodes_to_dicts(nodes_buffer)
            collection.add(
                ids=nodes_buffer_dict["ids"],
                documents=nodes_buffer_dict["texts"],
                metadatas=nodes_buffer_dict["metadatas"],
            )

        print(f"Finished loading {file_path}.")


def embed_non_pdf(files: list, user_id, collection):
    reader = UnstructuredReader()
    xlsx_reader = XlsxReader()
    non_pdf_documents = SimpleDirectoryReader(
        input_files=files,
        file_metadata=custom_metadata(user_id),
        required_exts=[".csv", ".jpg", ".xlsx"],
        file_extractor={
            ".csv": reader,
            ".jpg": reader,
            ".xlsx": xlsx_reader,
        },
    ).load_data(show_progress=True)

    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(chunk_size=1024, chunk_overlap=20),
        ]
    )

    non_pdf_nodes = pipeline.run(documents=non_pdf_documents)

    nodes_buffer = []
    for i, node in enumerate(non_pdf_nodes):
        node.node_id = str(uuid.uuid4())
        node.metadata["chunk_id"] = node.node_id

        nodes_buffer.append(node)
        if i % 25 == 0:
            nodes_buffer_dict = nodes_to_dicts(nodes_buffer)
            collection.add(
                ids=nodes_buffer_dict["ids"],
                documents=nodes_buffer_dict["texts"],
                metadatas=nodes_buffer_dict["metadatas"],
            )
            nodes_buffer = []

    if nodes_buffer:
        nodes_buffer_dict = nodes_to_dicts(nodes_buffer)
        collection.add(
            ids=nodes_buffer_dict["ids"],
            documents=nodes_buffer_dict["texts"],
            metadatas=nodes_buffer_dict["metadatas"],
        )

    return non_pdf_nodes


def update(data_dir, user_id):
    added_files, removed_files = read_changes(data_dir)

    chroma_db = chromadb.HttpClient(
        host=config.chroma_host, port=config.chroma_db_port
    )
    print("Chroma Setup done!")
    schema = IndexSchema.from_yaml(
        os.path.join(config.module_root_dir, "custom_schema.yaml")
    )
    redis_client = Redis(host=config.redis_host,port=6379,username="default",password=config.redis_password,db=0)
    vector_store = RedisVectorStore(
        redis_client=redis_client, schema=schema, overwrite=True
    )

    collection = chroma_db.get_or_create_collection(
        name=config.user_uploads_collection,
        embedding_function=HuggingFaceEmbeddingServer(
            url=config.tei_url + "/" + config.embedding + "/embed"
        ),
        metadata={
            "hnsw:batch_size": 512,
            "hnsw:sync_threshold": 1024,
        },
    )

    if len(removed_files) > 0:
        # TODO: create redis removal
        for file in removed_files:
            print(f"Removing: {file}")
            collection.delete(where={"file_path": file})
        print("Removal done.")

    elif len(added_files) > 0:
        total_nodes = []

        pdf_files = [file for file in added_files if file.endswith(".pdf")]

        non_pdf_files = list(set(added_files) - set(pdf_files))

        if len(non_pdf_files) > 0:
            try:
                non_pdf_nodes = embed_non_pdf(non_pdf_files, user_id, collection)
                total_nodes.extend(non_pdf_nodes)
            except:
                pass

        if len(pdf_files) > 0:
            try:
                pdf_nodes = embed_pdf(pdf_files, user_id, collection)
                total_nodes.extend(pdf_nodes)
            except:
                pass

        print("Chroma load Done!")

        # TODO: change index name
        try:
            load_redis(nodes=total_nodes, index_name="temka_testing")
        except:
            pass
        print("Redis load Done!")
    else:
        print("No changes to be done.")
    write_changes(data_dir, added_files, removed_files)
